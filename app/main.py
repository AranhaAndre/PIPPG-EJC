"""App de DOAÇÕES da Mocidade — isolado e autossuficiente (não é o ARES)."""
from __future__ import annotations

import asyncio
import io
import sqlite3
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path

from fastapi import Depends, FastAPI, HTTPException, Request, Response, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from .config import settings
from .database import Base, SessionLocal, engine, get_db
from .models import Category, Donation, Item
from .pix import pix_context, qr_svg
from .realtime import manager
from .schemas import (
    CategoryCreate,
    CategoryUpdate,
    DonationAdminUpdate,
    DonationCreate,
    DonationOut,
    ItemCreate,
    ItemUpdate,
    LoginIn,
)
from .security import COOKIE_NAME, check_credentials, create_token, require_admin
from .seed_data import CATEGORIAS, ITENS

WEB = Path(__file__).resolve().parent.parent / "web"
LOGO_IGREJA = WEB / "static" / "img" / "logo-igreja.png"
LOGO_EJC = WEB / "static" / "img" / "logo-ejc.png"


# ---------------------------------------------------------- backup ----
BACKUP_EVERY_HOURS = 6
BACKUP_KEEP = 12  # mantém as últimas 12 cópias (~3 dias de histórico)


def _db_file_path() -> Path | None:
    """Extrai o caminho do arquivo SQLite da DATABASE_URL."""
    url = settings.DATABASE_URL
    if url.startswith("sqlite") and "://" in url:
        p = "/" + url.split("://", 1)[1].lstrip("/")
        return Path(p)
    return None


def _do_backup() -> Path | None:
    """Cópia consistente do banco (usa a API de backup do SQLite)."""
    src = _db_file_path()
    if not src or not src.exists():
        return None
    bkdir = src.parent / "backups"
    bkdir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now()
    dest = bkdir / f"doacoes-{ts:%Y%m%d-%H%M%S}.db"
    s = sqlite3.connect(str(src))
    d = sqlite3.connect(str(dest))
    try:
        with d:
            s.backup(d)
    finally:
        s.close()
        d.close()
    # remove as cópias mais antigas além do limite
    olds = sorted(bkdir.glob("doacoes-*.db"))
    for o in olds[:-BACKUP_KEEP]:
        try:
            o.unlink()
        except OSError:
            pass
    (src.parent / "last_backup.txt").write_text(ts.isoformat())
    return dest


def _backup_info() -> dict:
    src = _db_file_path()
    info = {"last": None, "count": 0, "every_hours": BACKUP_EVERY_HOURS}
    if not src:
        return info
    marker = src.parent / "last_backup.txt"
    if marker.exists():
        try:
            info["last"] = marker.read_text().strip()
        except OSError:
            pass
    bkdir = src.parent / "backups"
    if bkdir.exists():
        info["count"] = len(list(bkdir.glob("doacoes-*.db")))
    return info


async def _backup_loop():
    await asyncio.sleep(20)  # primeiro backup logo após subir
    while True:
        try:
            await asyncio.to_thread(_do_backup)
        except Exception:
            pass
        await asyncio.sleep(BACKUP_EVERY_HOURS * 3600)


# ---------------------------------------------------------------- startup ----
@asynccontextmanager
async def lifespan(app: FastAPI):
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    await _seed_if_empty()
    backup_task = asyncio.create_task(_backup_loop())
    yield
    backup_task.cancel()
    await engine.dispose()


async def _seed_if_empty() -> None:
    async with SessionLocal() as db:
        existing = (await db.execute(select(Item.id).limit(1))).first()
        if existing:
            return
        cat_by_name: dict[str, Category] = {}
        for nome, ordem in CATEGORIAS:
            c = Category(nome=nome, ordem=ordem)
            db.add(c)
            cat_by_name[nome] = c
        await db.flush()  # gera os ids das categorias
        for ordem, (nome, meta, unidade, cat) in enumerate(ITENS):
            db.add(Item(
                nome=nome, meta=meta, unidade=unidade,
                category_id=cat_by_name[cat].id if cat in cat_by_name else None,
                ordem=ordem,
            ))
        await db.commit()


app = FastAPI(title="Doações · Mocidade", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=WEB / "static"), name="static")


# ---------------------------------------------------------- lógica comum ----
async def compute_progress(db: AsyncSession) -> dict:
    itens = (
        await db.execute(select(Item).where(Item.ativo == True).order_by(Item.ordem))  # noqa: E712
    ).scalars().all()
    cats = {c.id: c for c in (await db.execute(select(Category))).scalars().all()}

    doacoes = (
        await db.execute(select(Donation).where(Donation.status != "cancelado"))
    ).scalars().all()

    por_item: dict[int, dict] = {}
    for d in doacoes:
        if d.item_id is None:
            continue
        bucket = por_item.setdefault(d.item_id, {"doado": 0.0, "recebido": 0.0, "nomes": []})
        bucket["doado"] += d.quantidade
        if d.status == "recebido":
            bucket["recebido"] += d.quantidade
        nome_ex = d.doador_nome + (f" ({d.grupo})" if d.grupo else "")
        if nome_ex not in bucket["nomes"]:
            bucket["nomes"].append(nome_ex)

    itens_out = []
    completos = 0
    for it in itens:
        b = por_item.get(it.id, {"doado": 0.0, "recebido": 0.0, "nomes": []})
        pct = (b["doado"] / it.meta * 100) if it.meta else 0
        completo = b["doado"] >= it.meta
        completos += 1 if completo else 0
        cat = cats.get(it.category_id)
        itens_out.append({
            "id": it.id,
            "nome": it.nome,
            "category_id": it.category_id,
            "categoria": cat.nome if cat else "Sem categoria",
            "cat_ordem": cat.ordem if cat else 999,
            "meta": it.meta,
            "unidade": it.unidade,
            "doado": round(b["doado"], 2),
            "recebido": round(b["recebido"], 2),
            "percentual": round(pct, 1),
            "doadores": b["nomes"],
            "completo": completo,
        })

    total = len(itens_out)
    stats = {
        "itens_total": total,
        "itens_completos": completos,
        "cobertura_pct": round(completos / total * 100, 1) if total else 0,
        "doacoes_total": len(doacoes),
        "doadores_unicos": len({d.doador_nome for d in doacoes}),
    }
    return {"itens": itens_out, "stats": stats}


async def broadcast_progress(db: AsyncSession) -> None:
    snap = await compute_progress(db)
    await manager.broadcast({"type": "progress", **snap})


# -------------------------------------------------------------- páginas ----
@app.get("/", include_in_schema=False)
async def page_index():
    return FileResponse(WEB / "index.html")


@app.get("/admin", include_in_schema=False)
async def page_admin():
    return FileResponse(WEB / "admin.html")


@app.get("/login", include_in_schema=False)
async def page_login():
    return FileResponse(WEB / "login.html")


# ------------------------------------------------------ API pública ----
@app.get("/api/config")
async def api_config():
    return {
        "event_name": settings.EVENT_NAME,
        "event_subtitle": settings.EVENT_SUBTITLE,
        "event_verse": settings.EVENT_VERSE,
        "event_verse_ref": settings.EVENT_VERSE_REF,
        "pix": pix_context(),
    }


@app.get("/api/pix-qr.svg", include_in_schema=False)
async def api_pix_qr():
    svg = qr_svg()
    if not svg:
        raise HTTPException(404, "PIX não configurado")
    return Response(content=svg, media_type="image/svg+xml")


@app.get("/api/progress")
async def api_progress(db: AsyncSession = Depends(get_db)):
    return await compute_progress(db)


@app.post("/api/donations", response_model=DonationOut)
async def api_create_donation(payload: DonationCreate, db: AsyncSession = Depends(get_db)):
    # honeypot: bots preenchem "website"
    if payload.website:
        raise HTTPException(400, "Requisição inválida")

    if payload.item_id is None and not (payload.item_livre and payload.item_livre.strip()):
        raise HTTPException(400, "Escolha um item da lista ou descreva o que vai doar")

    unidade = payload.unidade or ""
    if payload.item_id is not None:
        item = await db.get(Item, payload.item_id)
        if not item or not item.ativo:
            raise HTTPException(404, "Item não encontrado")
        unidade = unidade or item.unidade

        # trava: não pode ultrapassar a meta do item
        total_doado = (await db.execute(
            select(func.coalesce(func.sum(Donation.quantidade), 0.0))
            .where(Donation.item_id == item.id, Donation.status != "cancelado")
        )).scalar() or 0.0
        faltante = round(item.meta - total_doado, 4)
        EPS = 1e-6
        if faltante <= EPS:
            raise HTTPException(400, f"“{item.nome}” já está completo. Obrigado! 🎯")
        if payload.quantidade > faltante + EPS:
            falta_txt = f"{faltante:g} {item.unidade}".strip()
            raise HTTPException(
                400,
                f"Faltam apenas {falta_txt} de “{item.nome}”. "
                f"Registre no máximo essa quantidade.",
            )

    d = Donation(
        item_id=payload.item_id,
        item_livre=(payload.item_livre or "").strip() or None,
        doador_nome=payload.doador_nome.strip(),
        grupo=(payload.grupo or "").strip() or None,
        contato=(payload.contato or "").strip() or None,
        quantidade=payload.quantidade,
        unidade=unidade,
        observacao=(payload.observacao or "").strip() or None,
        status="prometido",
    )
    db.add(d)
    await db.commit()
    await db.refresh(d)
    await broadcast_progress(db)
    return d


# ------------------------------------------------------------ auth ----
@app.post("/api/login")
async def api_login(body: LoginIn, request: Request):
    if not check_credentials(body.usuario, body.senha):
        raise HTTPException(401, "Usuário ou senha incorretos")
    token = create_token()
    resp = JSONResponse({"ok": True})
    # Secure só quando a conexão é HTTPS (evita o cookie "sumir" em HTTP e
    # causar o loop de login). Com --proxy-headers, o scheme reflete o proxy.
    resp.set_cookie(
        COOKIE_NAME, token,
        httponly=True, samesite="lax", secure=(request.url.scheme == "https"),
        max_age=settings.TOKEN_TTL_HOURS * 3600, path="/",
    )
    return resp


@app.post("/api/logout")
async def api_logout():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie(COOKIE_NAME, path="/")
    return resp


@app.get("/api/me")
async def api_me(admin: str = Depends(require_admin)):
    return {"usuario": admin}


# ------------------------------------------------------ API admin: doações ----
@app.get("/api/admin/donations")
async def admin_list(
    status: str | None = None,
    admin: str = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Donation).order_by(Donation.criado_em.desc())
    if status:
        stmt = stmt.where(Donation.status == status)
    rows = (await db.execute(stmt)).scalars().all()
    items = {i.id: i.nome for i in (await db.execute(select(Item))).scalars().all()}
    out = []
    for d in rows:
        out.append({
            **DonationOut.model_validate(d).model_dump(),
            "item_nome": items.get(d.item_id) if d.item_id else d.item_livre,
        })
    return out


@app.patch("/api/admin/donations/{donation_id}")
async def admin_update(
    donation_id: int,
    body: DonationAdminUpdate,
    admin: str = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    d = await db.get(Donation, donation_id)
    if not d:
        raise HTTPException(404, "Doação não encontrada")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(d, field, value)
    await db.commit()
    await broadcast_progress(db)
    return {"ok": True}


@app.delete("/api/admin/donations/{donation_id}")
async def admin_delete(
    donation_id: int,
    admin: str = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    d = await db.get(Donation, donation_id)
    if not d:
        raise HTTPException(404, "Doação não encontrada")
    await db.delete(d)
    await db.commit()
    await broadcast_progress(db)
    return {"ok": True}


# ------------------------------------------------------ API admin: categorias ----
@app.get("/api/admin/categories")
async def admin_categories(admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    cats = (await db.execute(select(Category).order_by(Category.ordem))).scalars().all()
    # contagem de itens por categoria
    counts = dict(
        (await db.execute(
            select(Item.category_id, func.count(Item.id)).group_by(Item.category_id)
        )).all()
    )
    return [
        {"id": c.id, "nome": c.nome, "ordem": c.ordem, "ativo": c.ativo,
         "itens": int(counts.get(c.id, 0))}
        for c in cats
    ]


@app.post("/api/admin/categories")
async def admin_create_category(
    body: CategoryCreate, admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db)
):
    ordem = body.ordem
    if ordem is None:
        mx = (await db.execute(select(func.max(Category.ordem)))).scalar() or 0
        ordem = mx + 1
    c = Category(nome=body.nome.strip(), ordem=ordem)
    db.add(c)
    await db.commit()
    await db.refresh(c)
    await broadcast_progress(db)
    return {"id": c.id, "nome": c.nome, "ordem": c.ordem}


@app.patch("/api/admin/categories/{cat_id}")
async def admin_update_category(
    cat_id: int, body: CategoryUpdate,
    admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db),
):
    c = await db.get(Category, cat_id)
    if not c:
        raise HTTPException(404, "Categoria não encontrada")
    data = body.model_dump(exclude_none=True)
    if "nome" in data:
        data["nome"] = data["nome"].strip()
    for f, v in data.items():
        setattr(c, f, v)
    await db.commit()
    await broadcast_progress(db)
    return {"ok": True}


@app.delete("/api/admin/categories/{cat_id}")
async def admin_delete_category(
    cat_id: int, admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db),
):
    c = await db.get(Category, cat_id)
    if not c:
        raise HTTPException(404, "Categoria não encontrada")
    # itens da categoria ficam sem categoria (ondelete=SET NULL cuida no banco,
    # mas garantimos aqui para o snapshot ficar coerente na hora)
    itens = (await db.execute(select(Item).where(Item.category_id == cat_id))).scalars().all()
    for it in itens:
        it.category_id = None
    await db.delete(c)
    await db.commit()
    await broadcast_progress(db)
    return {"ok": True}


# ------------------------------------------------------ API admin: itens ----
@app.post("/api/admin/items")
async def admin_create_item(
    body: ItemCreate, admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db),
):
    category_id = body.category_id
    # cria categoria na hora, se pedido
    if body.categoria_nova and body.categoria_nova.strip():
        mx = (await db.execute(select(func.max(Category.ordem)))).scalar() or 0
        nova = Category(nome=body.categoria_nova.strip(), ordem=mx + 1)
        db.add(nova)
        await db.flush()
        category_id = nova.id
    if category_id is not None and not await db.get(Category, category_id):
        raise HTTPException(404, "Categoria não encontrada")

    ordem = body.ordem
    if ordem is None:
        mx = (await db.execute(select(func.max(Item.ordem)))).scalar() or 0
        ordem = mx + 1
    it = Item(
        nome=body.nome.strip(), meta=body.meta, unidade=body.unidade.strip(),
        category_id=category_id, ordem=ordem,
    )
    db.add(it)
    await db.commit()
    await db.refresh(it)
    await broadcast_progress(db)
    return {"id": it.id, "nome": it.nome}


@app.patch("/api/admin/items/{item_id}")
async def admin_update_item(
    item_id: int, body: ItemUpdate,
    admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db),
):
    it = await db.get(Item, item_id)
    if not it:
        raise HTTPException(404, "Item não encontrado")
    data = body.model_dump(exclude_none=True)
    if data.get("category_id") is not None and not await db.get(Category, data["category_id"]):
        raise HTTPException(404, "Categoria não encontrada")
    for f, v in data.items():
        setattr(it, f, v.strip() if isinstance(v, str) else v)
    await db.commit()
    await broadcast_progress(db)
    return {"ok": True}


@app.post("/api/admin/items/{item_id}/reset")
async def admin_reset_item(
    item_id: int, admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db),
):
    """Apaga TODAS as doações do item, voltando-o a 'faltando tudo'."""
    it = await db.get(Item, item_id)
    if not it:
        raise HTTPException(404, "Item não encontrado")
    doacoes = (await db.execute(select(Donation).where(Donation.item_id == item_id))).scalars().all()
    n = len(doacoes)
    for d in doacoes:
        await db.delete(d)
    await db.commit()
    await broadcast_progress(db)
    return {"ok": True, "removidas": n}


@app.delete("/api/admin/items/{item_id}")
async def admin_delete_item(
    item_id: int, admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db),
):
    """Remove o item da lista (e suas doações, via cascade)."""
    it = await db.get(Item, item_id)
    if not it:
        raise HTTPException(404, "Item não encontrado")
    await db.delete(it)
    await db.commit()
    await broadcast_progress(db)
    return {"ok": True}


# ------------------------------------------------------ API admin: stats ----
@app.get("/api/admin/stats")
async def admin_stats(admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    snap = await compute_progress(db)
    # ranking "quem já entregou": só doações RECEBIDAS (confirmadas)
    rows = (await db.execute(select(Donation).where(Donation.status == "recebido"))).scalars().all()
    ranking: dict[str, int] = {}
    for d in rows:
        chave = d.doador_nome + (f" ({d.grupo})" if d.grupo else "")
        ranking[chave] = ranking.get(chave, 0) + 1
    top = sorted(ranking.items(), key=lambda kv: kv[1], reverse=True)[:10]
    por_cat: dict[str, dict] = {}
    for it in snap["itens"]:
        c = por_cat.setdefault(it["categoria"], {"total": 0, "completos": 0, "ordem": it["cat_ordem"]})
        c["total"] += 1
        c["completos"] += 1 if it["completo"] else 0
    cats_sorted = sorted(por_cat.items(), key=lambda kv: kv[1]["ordem"])
    return {
        **snap["stats"],
        "ranking": [{"nome": n, "doacoes": q} for n, q in top],
        "por_categoria": [
            {"categoria": k, "total": v["total"], "completos": v["completos"]}
            for k, v in cats_sorted
        ],
        "itens": snap["itens"],
    }


@app.get("/api/admin/backup-info")
async def admin_backup_info(admin: str = Depends(require_admin)):
    return _backup_info()


# ------------------------------------------------------ exportações ----
async def _export_rows(db: AsyncSession):
    rows = (await db.execute(select(Donation).order_by(Donation.criado_em))).scalars().all()
    items = {i.id: i.nome for i in (await db.execute(select(Item))).scalars().all()}
    return rows, items


@app.get("/api/admin/export.xlsx", include_in_schema=False)
async def admin_export_xlsx(admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.drawing.image import Image as XLImage

    rows, items = await _export_rows(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Doações"

    # ---- cabeçalho padronizado com logos (igreja à esquerda, EJC à direita) ----
    ws.row_dimensions[1].height = 46
    ws.merge_cells("B1:I1")
    ws.merge_cells("B2:I2")
    c1 = ws["B1"]; c1.value = settings.EVENT_NAME
    c1.font = Font(name="Calibri", size=16, bold=True, color="10214B")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c2 = ws["B2"]; c2.value = settings.EVENT_SUBTITLE
    c2.font = Font(name="Calibri", size=10, color="5A5A5A")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    try:
        if LOGO_IGREJA.exists():
            ig = XLImage(str(LOGO_IGREJA)); ig.width = 46; ig.height = 46
            ws.add_image(ig, "A1")
    except Exception:
        pass
    try:
        if LOGO_EJC.exists():
            ej = XLImage(str(LOGO_EJC)); ej.height = 40; ej.width = int(40 * 213 / 120)
            ws.add_image(ej, "J1")
    except Exception:
        pass

    # ---- tabela (linha 4 em diante) ----
    header = ["ID", "Item", "Doador", "Grupo", "Contato", "Quantidade", "Unidade",
              "Status", "Observação", "Registrado em"]
    ws.append([])  # linha 3 em branco
    ws.append(header)
    head_fill = PatternFill("solid", fgColor="10214B")
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = head_fill
    for d in rows:
        ws.append([
            d.id,
            items.get(d.item_id) if d.item_id else (d.item_livre or ""),
            d.doador_nome, d.grupo or "", d.contato or "", d.quantidade, d.unidade,
            d.status, d.observacao or "",
            d.criado_em.strftime("%d/%m/%Y %H:%M") if d.criado_em else "",
        ])
    widths = [6, 26, 20, 14, 16, 12, 12, 12, 26, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=doacoes.xlsx"},
    )


@app.get("/api/admin/export.pdf", include_in_schema=False)
async def admin_export_pdf(admin: str = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    from fpdf import FPDF

    snap = await compute_progress(db)
    st = snap["stats"]

    def txt(s: str) -> str:
        # fpdf core fonts usam latin-1; normaliza o que não couber
        return (s or "").encode("latin-1", "replace").decode("latin-1")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(left=15, top=12, right=15)
    pdf.add_page()

    # ---- cabeçalho padronizado: logo igreja à esquerda, EJC à direita ----
    try:
        if LOGO_IGREJA.exists():
            pdf.image(str(LOGO_IGREJA), x=15, y=11, w=20, h=20)
    except Exception:
        pass
    try:
        if LOGO_EJC.exists():
            # proporção 213:120 -> largura ~30 para altura 17
            pdf.image(str(LOGO_EJC), x=195 - 30, y=13, w=30, h=17)
    except Exception:
        pass

    pdf.set_xy(15, 13)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(16, 33, 75)
    pdf.cell(0, 9, txt(settings.EVENT_NAME), ln=1, align="C")
    pdf.set_x(15)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(90, 90, 90)
    pdf.cell(0, 6, txt(settings.EVENT_SUBTITLE), ln=1, align="C")
    pdf.set_x(15)
    pdf.cell(0, 6, txt("Relatório gerado em " + datetime.now().strftime("%d/%m/%Y %H:%M")), ln=1, align="C")
    pdf.set_y(35)
    pdf.set_draw_color(224, 230, 245)
    pdf.line(15, 34, 195, 34)
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(20, 20, 20)
    pdf.cell(0, 7, txt(
        f"Cobertura: {st['cobertura_pct']}%  |  "
        f"Itens completos: {st['itens_completos']}/{st['itens_total']}  |  "
        f"Doações: {st['doacoes_total']}  |  Pessoas: {st['doadores_unicos']}"
    ), ln=1)
    pdf.ln(3)

    # agrupa por categoria (respeitando a ordem)
    por_cat: dict = {}
    for it in snap["itens"]:
        por_cat.setdefault((it["cat_ordem"], it["categoria"]), []).append(it)

    for (_, catnome), itens in sorted(por_cat.items(), key=lambda kv: kv[0][0]):
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(238, 242, 251)
        pdf.set_text_color(16, 33, 75)
        pdf.cell(0, 8, txt(catnome), ln=1, fill=True)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(110, 110, 110)
        pdf.cell(70, 6, txt("Item"))
        pdf.cell(28, 6, txt("Meta"))
        pdf.cell(22, 6, txt("Status"))
        pdf.cell(0, 6, txt("Doadores"), ln=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(20, 20, 20)
        for it in itens:
            status = "OK" if it["completo"] else f"{it['percentual']:.0f}%"
            doadores = ", ".join(it["doadores"]) if it["doadores"] else "-"
            if len(doadores) > 58:
                doadores = doadores[:57] + "…"
            pdf.cell(70, 6, txt(it["nome"][:38]))
            pdf.cell(28, 6, txt(f"{it['meta']:g} {it['unidade']}"[:16]))
            pdf.cell(22, 6, txt(status))
            pdf.cell(0, 6, txt(doadores), ln=1)
        pdf.ln(2)

    out = pdf.output()
    buf = io.BytesIO(bytes(out))
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=relatorio_doacoes.pdf"},
    )


# ------------------------------------------------------ websocket ----
@app.websocket("/ws")
async def ws_endpoint(ws: WebSocket):
    await manager.connect(ws)
    try:
        async with SessionLocal() as db:
            snap = await compute_progress(db)
        await ws.send_json({"type": "progress", **snap})
        while True:
            await ws.receive_text()  # mantém a conexão viva (ping do cliente)
    except WebSocketDisconnect:
        await manager.disconnect(ws)
    except Exception:
        await manager.disconnect(ws)


@app.get("/health", include_in_schema=False)
async def health():
    return {"status": "ok"}
